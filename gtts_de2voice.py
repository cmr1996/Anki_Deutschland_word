from gtts import gTTS
import openpyxl
from openpyxl import load_workbook

# 元音替换
semivowel = {'a': 'ä', 'o': 'ö', 'u': 'ü'}
# 类型判断哈希
IsRightType = {'Adj.':1 , 'Adv.':1, 'P.II':1, 'Konj.':1}


def insertStr(word, origin_char, insert_char):
    # 把字符串转为 list
    str_list = list(word)
    # 字符数， 可以利用这个在某个位置插入字符
    #count = len(str_list)
    # 找到 斜杠的位置
    nPos = str_list.index(origin_char)
    # 在斜杠位置之前 插入要插入的字符
    str_list.insert(nPos, insert_char)
    # 将 list 转为 str
    return "".join(str_list)


# 变音
def tailchange(word):
    for i in range(len(word)-1, -1, -1):
        if word[i] == 'a' or word[i] == 'o':
            word = word[:i] + semivowel[word[i]] + word[i+1:]
            # word[i] = semivowel[word[i]]
        elif word[i] == 'u':
            if word[i-1] == 'a':
                word = word[:i-1] + semivowel[word[i-1]] + word[i:]
                # word[i-1] = semivowel[word[i-1]]
            else:
                word = word[:i] + semivowel[word[i]] + word[i+1:]
                # word[i] = semivowel[word[i]]
     

# 名词处理 
# 例'die', 'Linie, -n'
# 例'das', 'Rathaus, .. er'
def noun_handle(gender, word):
    # 中文标点变英文
    word = word.replace("，",",")
    word_tmp = word.split(',')
    noun = word_tmp[0]
    # 无复数形式名词特殊处理
    if len(word_tmp) == 1:
        tail = '-'
    else:
        tail = word_tmp[1]

    # 判断是否变音
    index = tail.find('..')
    if index != -1: # 变音
        tail = tail[index+2:]
        tail = tail.strip() # 去除头尾空格
        tailchange(noun)
    #elif tail.find('..') != -1: # 变音
    else: # 非变音
        index = tail.find('-')    
        tail = tail[index+1:]
        tail = tail.strip() # 去除头尾空格

    # 保存
    nameforsave = noun
    download_name = gender + ' ' + noun + '.' + '  die ' + noun + tail
    return download_name, nameforsave


# 非名词处理
def imnoun_handle(word):
    # 判断类型
    word_tmp = word.split()
    word_count = len(word_tmp)
    Type = word_tmp[word_count - 1]
    # 动词
    if Type == 'Vi.' or Type == 'Vr.' or Type == 'Vt.':
        word = word.replace(',', '.')
        word = word.replace('，', '.')
        word = word.replace('（', '(')
        word = insertStr(word, '(', '.')
        download_name = word.rstrip(Type)
        nameforsave = word_tmp[0] + word_tmp[1]
    # 非动词
    else:    
        if word_count > 1: # 超过一个单词
            if IsRightType[Type]:  # 如果单词是adj/adv/第二分词/第一分词
                nameforsave = word_tmp[0] + word_tmp[1]
                download_name = word.rstrip(Type)
            else:
                download_name = word
                nameforsave = download_name
        else:  # 仅有一个单词
            nameforsave = word
            download_name = nameforsave

    return download_name, nameforsave


# 字符串处理
def word_handle(gender, word):
    #  返回下载字符串download_name和保存字符串nameforsave
    if not gender:
        download_name, nameforsave = imnoun_handle(word)
    else:
        download_name, nameforsave = noun_handle(gender, word)
    
    # 下载和保存
    sound_name = '[sound:'+ nameforsave + '.mp3]'
    tts = gTTS(download_name, lang='de')
    # save_path = 'E:\\test\\word_anki\\DE_word_voice\\anki_sound\\' + sound_name + '.mp3'
    # tts.save(save_path)
    nameforsave += '.mp3'
    tts.save(nameforsave)
    print("音频下载成功:", sound_name)
    return sound_name


# excel操作
def excel_handle(file):
    # 初始化
    wb = load_workbook(filename=file)
    ws = wb.active

    # 循环Excel文件的所有行
    rows_total = len(list(ws.rows))
    for i in range(rows_total):
        row_index = i + 1
        gender = ws['A' + str(row_index)].value
        word = ws['B'+ str(row_index) ].value
        Is_hanlde = ws['F' + str(row_index)].value
        if Is_hanlde:
            print('单词', word, '已经完成')
            continue
        elif not word: # 跳过空行
            continue
        # chinese = sheet['C' + str(row_index)]
        # example = sheet['D' + str(row_index)]
        # 一下为字符串处理,
        # 字符串处理返回一个单词用于回写入excel
        sound_name = word_handle(gender, word)
        # sound_name = 0
        ws['E'+ str(row_index)] = sound_name
        ws['F' + str(row_index)] = 1
        # 结束操作，保存excel
        wb.save(filename=file)
        print('成功回写:', sound_name)
    



# 主函数
if __name__ == '__main__':
    filename = 'lektion15.xlsx'
    excel_handle(filename)

# 典型动词字符串
# 动词短语:sich die Zähne putzen
# 普通动词:entlanggehen (geht entlang, ging entlang, ist entlanggegangen) Vi.
# 反身动词: sich etw. (A) ansehen (sieht sich an, sah sich an, hat sich angesehen) Vr.


# 典型名词字符串
# 普通名词: die	Linie, -n
# 带变音名词: das	Rathaus, .. er
# 非正常的名词变化:  kaufman --> kaufleute