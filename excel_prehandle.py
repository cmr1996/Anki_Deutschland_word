import openpyxl
from openpyxl import load_workbook
from queue import Queue,LifoQueue,PriorityQueue

# 变量
filename = 'lektion18.xlsx'
delete_line_index = Queue(maxsize=0)

# 初始化
wb = load_workbook(filename)
ws = wb.active

# 循环Excel文件的所有行
rows_total = len(list(ws.rows))

for i in range(rows_total):
    row_index = i + 1
    gender = ws['A' + str(row_index)].value
    word = ws['B'+ str(row_index) ].value
    chinese = ws['C'+ str(row_index) ].value
    
    if gender == '_____________________________________________________________________':
        ws['A' + str(row_index)] = ''
    elif gender:
        gender_tmp = gender.split()
        word_cnt = len(gender_tmp)
        if word_cnt > 2:
            ws['A' + str(row_index)] = ''
            ws['D' + str(row_index - 1)] = gender

    # 结束操作，保存excel
    wb.save(filename)

for i in range(rows_total):
    row_index = i + 1
    gender = ws['A' + str(row_index)].value
    word = ws['B'+ str(row_index) ].value

    if not gender and not word:
        delete_line_index.put(row_index)
    elif not delete_line_index.empty():
        dst_row_index = delete_line_index.get()
        ws['A' + str(dst_row_index)] = ws['A' + str(row_index)].value
        ws['B' + str(dst_row_index)] = ws['B' + str(row_index)].value
        ws['C' + str(dst_row_index)] = ws['C' + str(row_index)].value
        ws['D' + str(dst_row_index)] = ws['D' + str(row_index)].value
        delete_line_index.put(row_index)

    # 结束操作，保存excel
    wb.save(filename)
